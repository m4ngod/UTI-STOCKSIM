import os
import pytest
from app.services.export_service import ExportService, ExportServiceError
import app.services.export_service as es_mod
from observability.metrics import metrics


def test_export_excel_with_meta(tmp_path):
    es = ExportService()
    meta = {"snapshot_id": "sid_meta", "source": "unit", "extra": 123}
    data = [
        {"a": 1, "b": "x", "snapshot_id": "sid_meta"},
        {"a": 2, "b": "y", "snapshot_id": "sid_meta"},
    ]
    base_excel = metrics.counters.get("export_excel_success", 0)
    try:
        path = es.export("xlsx", data, meta, file_path=tmp_path / "out_excel")
    except ExportServiceError as e:
        # 缺失 pandas/openpyxl 或 xlsxwriter 引擎 -> 允许直接通过
        if e.code == 'PANDAS_MISSING':
            return
        raise
    # 成功路径校验
    assert os.path.isfile(path)
    assert path.endswith('.xlsx')
    assert metrics.counters.get("export_excel_success", 0) >= base_excel + 1
    # 若安装 openpyxl, 进一步校验 META sheet
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:  # 未安装则跳过内容校验
        return
    wb = load_workbook(path, read_only=True)
    assert 'META' in wb.sheetnames
    meta_ws = wb['META']
    headers = [c.value for c in next(meta_ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ['key', 'value']
    rows = {r[0].value: r[1].value for r in meta_ws.iter_rows(min_row=2)}
    assert rows.get('snapshot_id') == 'sid_meta'
    assert rows.get('source') == 'unit'
    assert rows.get('extra') == 123


def test_export_excel_openpyxl_fallback(tmp_path):
    if es_mod._OpenpyxlWorkbook is None:  # type: ignore[attr-defined]
        pytest.skip('openpyxl not installed; skip fallback test')
    es = ExportService()
    meta = {"snapshot_id": "sid_fb", "source": "fb"}
    data = [{"c": 10, "snapshot_id": "sid_fb"}]
    orig_pd = es_mod._pd
    es_mod._pd = None  # type: ignore
    try:
        path = es.export('excel', data, meta, file_path=tmp_path / 'fallback')
        assert os.path.isfile(path)
        assert path.endswith('.xlsx')
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(path, read_only=True)
        assert 'META' in wb.sheetnames and 'DATA' in wb.sheetnames
        meta_ws = wb['META']
        rows = {r[0].value: r[1].value for r in meta_ws.iter_rows(min_row=2)}
        assert rows.get('snapshot_id') == 'sid_fb'
        assert rows.get('source') == 'fb'
    finally:
        es_mod._pd = orig_pd  # 恢复
